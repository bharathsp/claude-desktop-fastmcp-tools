"""Thread-safe Excel session logger."""

from __future__ import annotations

import json
import threading
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from session_logger.config import resolve_log_path, settings

HEADERS = ["Timestamp", "User IP Address", "Host", "Question", "Response"]

_lock = threading.Lock()


def _truncate(value: str) -> str:
    if len(value) <= settings.max_cell_length:
        return value
    return value[: settings.max_cell_length - 3] + "..."


def _get_sheet(path) -> tuple[Any, Worksheet]:
    if path.exists():
        workbook = load_workbook(path)
        sheet = workbook.active
        if sheet.max_row == 0 or sheet.cell(1, 1).value != HEADERS[0]:
            sheet.delete_rows(1, sheet.max_row)
            sheet.append(HEADERS)
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Session Logs"
        sheet.append(HEADERS)
    return workbook, sheet


def log_session(
    *,
    user_ip: str,
    host: str,
    question: str,
    response: str,
) -> None:
    """Append one session row to the Excel log file."""
    if not settings.enabled:
        return

    timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    row = [
        timestamp,
        _truncate(str(user_ip)),
        _truncate(str(host)),
        _truncate(str(question)),
        _truncate(str(response)),
    ]

    path = resolve_log_path()

    with _lock:
        workbook, sheet = _get_sheet(path)
        sheet.append(row)
        workbook.save(path)


def format_question(tool_or_action: str, payload: Any) -> str:
    return json.dumps({"action": tool_or_action, "input": payload}, default=str)


def format_response(payload: Any) -> str:
    if isinstance(payload, str):
        return payload
    return json.dumps(payload, default=str)
